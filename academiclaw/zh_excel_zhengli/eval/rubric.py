"""
EXCEL汇总表整理 — 评分脚本 (rubric.py)

任务概述:
    对学生奖学金信息汇总表进行数据清洗、去重、排序与排版美化，
    产出 EXCEL 汇总表整理_完成.xlsx，含四个 Sheet:
        汇总表（主表）、打印版、Sheet3（统计）、Sheet4（学院顺序）

总分 100 分，分项:
    F. 文件完整性      5 分  — 文件名、可打开、工作表齐全
    A. 去重与排序     25 分  — 去重、排序、序号、重复标红
    B. 数据规范化     20 分  — 学号/电话文本格式、金额、学位、生源地、民族、学院
    C. 格式与表格设置  30 分  — 标题、表头、字体、行高、边框、隐藏列、冻结
    D. 公式与函数     10 分  — LEN/IF 公式、Sheet3 COUNTIF/SUM
    E. Sheet2/Sheet3  10 分  — 打印版列/设置、统计表结构
"""

import os
import re
from typing import Tuple, Dict, Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_to_tuple
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _s(v) -> str:
    """Safely convert value to stripped string."""
    if v is None:
        return ""
    return str(v).strip()


def _cell_fill_rgb(cell) -> str:
    """Return the hex RGB of a cell's solid fill, or empty string."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return ""
        rgb = ""
        if fill.start_color and fill.start_color.rgb:
            rgb = str(fill.start_color.rgb)
        if not rgb and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
        return rgb.upper()
    except Exception:
        return ""


def _is_red_fill(cell) -> bool:
    """Check whether a cell has a red background (various common red hex)."""
    rgb = _cell_fill_rgb(cell)
    if not rgb:
        return False
    # Common red patterns: FFFF0000, FF0000, 00FF0000
    return "FF0000" in rgb


def _is_yellow_fill(cell) -> bool:
    """Check whether a cell has a yellow background."""
    rgb = _cell_fill_rgb(cell)
    if not rgb:
        return False
    return "FFFF00" in rgb


def _parse_int_amount(v) -> int:
    """Parse a monetary value to int. Returns -1 on failure."""
    if v is None:
        return -1
    if isinstance(v, (int, float)):
        return int(v)
    s = _s(v).replace("元", "").replace(",", "").replace("，", "")
    if not s:
        return -1
    try:
        return int(float(s))
    except Exception:
        return -1


def _degree_sort_key(v) -> int:
    s = _s(v)
    if "本" in s:
        return 0
    if "硕" in s:
        return 1
    if "博" in s:
        return 2
    return 99


def _review_sort_key(v) -> int:
    s = _s(v)
    if "新" in s:
        return 0
    if "续" in s:
        return 1
    if "补" in s:
        return 2
    return 99


def _headers_from_row(ws, row: int):
    """Return (list_of_header_strings, {header: 1-based_col_index})."""
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(_s(ws.cell(row, c).value))
    # Trim trailing empty
    while headers and headers[-1] == "":
        headers.pop()
    hmap = {}
    for i, h in enumerate(headers):
        if h and h not in hmap:
            hmap[h] = i + 1
    return headers, hmap


def _find_last_data_row(ws, col_a: int, col_b: int) -> int:
    """Scan from bottom to find last row with data in either column."""
    for r in range(ws.max_row, 2, -1):
        if ws.cell(r, col_a).value is not None or ws.cell(r, col_b).value is not None:
            return r
    return 2


def _dedup_key(sid_val, name_val, phone_val):
    """Build a dedup key: student-ID first, fallback to name+phone."""
    sid = _s(sid_val)
    if sid:
        return ("SID", sid)
    return ("NP", _s(name_val), _s(phone_val))


def _parse_academy_list_from_cell(raw) -> list:
    """Parse a cell that might contain newline/comma-separated academy names."""
    s = _s(raw)
    if not s:
        return []
    for sep in ["\n", "\r"]:
        if sep in s:
            parts = [p.strip() for p in re.split(r"[\r\n]+", s) if p.strip()]
            if len(parts) >= 3:
                return parts
    for sep_pattern in ["、", r"[,，;；]+"]:
        parts = [p.strip() for p in re.split(sep_pattern, s) if p.strip()]
        if len(parts) >= 5:
            return parts
    return []


def _extract_academy_order(ws4) -> list:
    """Extract ordered academy names from Sheet4.

    Sheet4 may present academies as:
    - A single cell with newline-separated list (most common)
    - A column of cells
    - A row of cells
    """
    skip = {"学院", "学院名称", "学院顺序", "学院排序", "序号"}

    # Pass 1: look for a single cell with a long multi-line list
    for r in range(1, ws4.max_row + 1):
        for c in range(1, ws4.max_column + 1):
            v = ws4.cell(r, c).value
            if v is None:
                continue
            items = _parse_academy_list_from_cell(v)
            if len(items) >= 5:
                out, seen = [], set()
                for it in items:
                    if it in skip or it in seen:
                        continue
                    out.append(it)
                    seen.add(it)
                return out

    # Pass 2: choose best row or column by count of non-empty text cells
    row_counts, col_counts = {}, {}
    for r in range(1, ws4.max_row + 1):
        for c in range(1, ws4.max_column + 1):
            v = _s(ws4.cell(r, c).value)
            if not v or v in skip:
                continue
            row_counts[r] = row_counts.get(r, 0) + 1
            col_counts[c] = col_counts.get(c, 0) + 1

    best_row = max(row_counts, key=row_counts.get) if row_counts else None
    best_col = max(col_counts, key=col_counts.get) if col_counts else None
    best_r_cnt = row_counts.get(best_row, 0) if best_row else 0
    best_c_cnt = col_counts.get(best_col, 0) if best_col else 0

    order, seen = [], set()

    if best_row is not None and best_r_cnt >= best_c_cnt and best_r_cnt > 0:
        for c in range(1, ws4.max_column + 1):
            v = _s(ws4.cell(best_row, c).value)
            if v and v not in skip and v not in seen:
                order.append(v)
                seen.add(v)
        return order

    if best_col is not None and best_c_cnt > 0:
        for r in range(1, ws4.max_row + 1):
            v = _s(ws4.cell(r, best_col).value)
            if v and v not in skip and v not in seen:
                order.append(v)
                seen.add(v)
        return order

    # Fallback: row-major scan
    for r in range(1, ws4.max_row + 1):
        for c in range(1, ws4.max_column + 1):
            v = _s(ws4.cell(r, c).value)
            if v and v not in skip and v not in seen:
                order.append(v)
                seen.add(v)
    return order


def _locate_deliverable(answer_dir: str) -> str:
    """Find the deliverable xlsx file in answer_dir. Returns path or ''."""
    if not os.path.isdir(answer_dir):
        return ""
    exact = os.path.join(answer_dir, "EXCEL 汇总表整理_完成.xlsx")
    if os.path.exists(exact):
        return exact
    # Fuzzy fallback
    for f in os.listdir(answer_dir):
        if f.endswith("_完成.xlsx"):
            return os.path.join(answer_dir, f)
    for f in os.listdir(answer_dir):
        if f.endswith(".xlsx"):
            return os.path.join(answer_dir, f)
    return ""


def _locate_input_xlsx() -> str:
    """Find the original input xlsx from context/."""
    query_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    ctx = os.path.join(query_root, "context")
    if not os.path.isdir(ctx):
        return ""
    p = os.path.join(ctx, "EXCEL汇总表整理.xlsx")
    if os.path.exists(p):
        return p
    for f in os.listdir(ctx):
        if f.endswith(".xlsx") and "完成" not in f:
            return os.path.join(ctx, f)
    return ""


# ---------------------------------------------------------------------------
# Expected header order for 汇总表 (row 2)
# ---------------------------------------------------------------------------
EXPECTED_VISIBLE_HEADERS = [
    "序号", "奖励/资助年度", "学号", "姓名", "班级", "学院",
    "奖学金名称", "金额", "性别", "在读学位", "生源地", "民族",
    "成绩", "专业排名", "联系电话", "E-mail", "基本情况", "评审类型",
]

EXPECTED_ALL_HEADERS = [
    "序号", "奖励/资助年度", "学号", "学号长度", "原姓名", "姓名",
    "班级", "学院", "奖学金名称", "金额", "性别", "在读学位",
    "生源地", "民族", "成绩", "专业排名", "联系电话", "E-mail",
    "基本情况", "评审类型",
]


# ---------------------------------------------------------------------------
# Main evaluation
# ---------------------------------------------------------------------------

def evaluate(answer_dir: str) -> Tuple[int, Dict[str, Any]]:
    """
    Evaluate the agent's EXCEL汇总表整理 output.

    Args:
        answer_dir: absolute path to the agent output directory
                    (e.g. /path/to/query/gpt-5/attempt_1)

    Returns:
        (score, report) — score in [0, 100], report dict with per-section detail
    """
    if openpyxl is None:
        return 0, {"error": "openpyxl not installed; cannot evaluate"}

    sections = {
        "F_文件完整性": {"max": 5, "score": 0, "details": {}},
        "A_去重与排序": {"max": 25, "score": 0, "details": {}},
        "B_数据规范化": {"max": 20, "score": 0, "details": {}},
        "C_格式与表格设置": {"max": 30, "score": 0, "details": {}},
        "D_公式与函数": {"max": 10, "score": 0, "details": {}},
        "E_Sheet2_Sheet3": {"max": 10, "score": 0, "details": {}},
    }

    def add(section, key, pts, msg):
        sections[section]["score"] += pts
        sections[section]["details"][key] = f"{pts} — {msg}"

    # ======================================================================
    # F. 文件完整性 (5 分)
    # ======================================================================
    fpath = _locate_deliverable(answer_dir)
    if not fpath:
        add("F_文件完整性", "F0_not_found", 0, "deliverable xlsx not found")
        total = 0
        return total, _build_report(sections, total)

    fname = os.path.basename(fpath)
    if fname == "EXCEL 汇总表整理_完成.xlsx":
        add("F_文件完整性", "F1_filename", 3, "exact filename match")
    elif "_完成" in fname and fname.endswith(".xlsx"):
        add("F_文件完整性", "F1_filename", 1, f"partial match: {fname}")
    else:
        add("F_文件完整性", "F1_filename", 0, f"wrong filename: {fname}")

    try:
        wb = openpyxl.load_workbook(fpath)
    except Exception as exc:
        add("F_文件完整性", "F2_open", 0, f"cannot open workbook: {exc}")
        total = sections["F_文件完整性"]["score"]
        return total, _build_report(sections, total)

    required_sheets = {"汇总表", "打印版", "Sheet3", "Sheet4"}
    present = set(wb.sheetnames)
    missing = required_sheets - present
    if not missing:
        add("F_文件完整性", "F2_sheets", 2, "all 4 required sheets present")
    else:
        add("F_文件完整性", "F2_sheets", 0, f"missing sheets: {missing}")

    # Need 汇总表 to proceed with further evaluation
    if "汇总表" not in present:
        total = sum(s["score"] for s in sections.values())
        return total, _build_report(sections, total)

    ws = wb["汇总表"]
    headers, hmap = _headers_from_row(ws, 2)

    def col(name):
        return hmap.get(name)

    sid_col = col("学号")
    sidlen_col = col("学号长度")
    origname_col = col("原姓名")
    name_col = col("姓名")
    acad_col = col("学院")
    amt_col = col("金额")
    deg_col = col("在读学位")
    origin_col = col("生源地")
    eth_col = col("民族")
    phone_col = col("联系电话")
    basic_col = col("基本情况")
    review_col = col("评审类型")

    # Check all critical columns exist
    critical_cols = {
        "学号": sid_col, "学号长度": sidlen_col, "原姓名": origname_col,
        "姓名": name_col, "学院": acad_col, "金额": amt_col,
        "在读学位": deg_col, "生源地": origin_col, "民族": eth_col,
        "联系电话": phone_col, "基本情况": basic_col, "评审类型": review_col,
    }
    missing_cols = [k for k, v in critical_cols.items() if v is None]
    if missing_cols:
        add("C_格式与表格设置", "C0_missing_headers", 0,
            f"critical headers missing: {missing_cols} — cannot evaluate further")
        total = sum(s["score"] for s in sections.values())
        return total, _build_report(sections, total)

    last_row = _find_last_data_row(ws, sid_col, name_col)

    # Collect data rows (rows 3..last_row that have data)
    data_rows = []
    for r in range(3, last_row + 1):
        if _s(ws.cell(r, sid_col).value) or _s(ws.cell(r, name_col).value):
            data_rows.append(r)

    # Extract academy order from Sheet4
    academy_order = []
    if "Sheet4" in present:
        academy_order = _extract_academy_order(wb["Sheet4"])
    acad_set = set(academy_order)
    acad_rank = {a: i for i, a in enumerate(academy_order)}

    # ==================================================================
    # C. 格式与表格设置 (30 分)
    # ==================================================================

    # C1: header order (6 分)
    if headers == EXPECTED_ALL_HEADERS:
        add("C_格式与表格设置", "C1_header_order", 6, "exact header order match")
    else:
        # Check visible headers (skip hidden cols)
        visible = []
        for i, h in enumerate(headers):
            c_idx = i + 1
            letter = get_column_letter(c_idx)
            dim = ws.column_dimensions.get(letter)
            if dim and dim.hidden:
                continue
            if h:
                visible.append(h)
        if visible == EXPECTED_VISIBLE_HEADERS:
            add("C_格式与表格设置", "C1_header_order", 4,
                "visible headers correct, but full order differs")
        else:
            add("C_格式与表格设置", "C1_header_order", 0,
                f"header mismatch; got {headers[:5]}...")

    # C2: title row (row 1) — merged, 黑体 18号 加粗 居中 (6 分)
    c2 = 0
    title_cell = ws["A1"]
    merged_a1 = any(str(rng).startswith("A1:") for rng in ws.merged_cells.ranges)
    if merged_a1:
        c2 += 1
    font = title_cell.font
    if font and font.bold and float(font.size or 0) == 18.0:
        c2 += 2
    if font and font.name in ("黑体", "SimHei"):
        c2 += 1
    align = title_cell.alignment
    if align and align.horizontal == "center":
        c2 += 1
    # No border on title row (lenient — just give the point)
    c2 += 1
    c2 = min(c2, 6)
    add("C_格式与表格设置", "C2_title_format", c2, f"title formatting ({c2}/6)")

    # C3: header row font — 宋体 12号 加粗 居中 (3 分)
    hfont = ws.cell(2, 1).font
    c3 = 0
    if hfont and hfont.bold:
        c3 += 1
    if hfont and float(hfont.size or 0) == 12.0:
        c3 += 1
    halign = ws.cell(2, 1).alignment
    if halign and halign.horizontal == "center":
        c3 += 1
    add("C_格式与表格设置", "C3_header_font", c3, f"header font ({c3}/3)")

    # C4: body font — 宋体 11号, not bold/italic (5 分)
    body_ok = True
    sample_end = min(last_row + 1, 25)
    for r in range(3, sample_end):
        bf = ws.cell(r, 1).font
        if bf is None:
            continue
        if bf.name not in ("宋体", "SimSun") or float(bf.size or 0) != 11.0:
            body_ok = False
            break
        if bf.bold or bf.italic or bf.underline:
            body_ok = False
            break
    add("C_格式与表格设置", "C4_body_font", 5 if body_ok else 0,
        "body font OK" if body_ok else "body font mismatch in sample rows")

    # C5: row height = 25 (3 分)
    rh_ok = True
    for r in range(1, min(last_row + 1, 100)):
        h = ws.row_dimensions[r].height
        if h is not None and abs(float(h) - 25.0) > 0.5:
            rh_ok = False
            break
    add("C_格式与表格设置", "C5_row_height", 3 if rh_ok else 0,
        "row height 25 OK" if rh_ok else "row height != 25")

    # C6: hidden columns — 学号长度 and 原姓名 (3 分)
    hidden_sidlen = ws.column_dimensions[get_column_letter(sidlen_col)].hidden
    hidden_origname = ws.column_dimensions[get_column_letter(origname_col)].hidden
    c6 = 0
    if hidden_sidlen:
        c6 += 1
    if hidden_origname:
        c6 += 2
    add("C_格式与表格设置", "C6_hidden_cols", c6,
        f"学号长度={'hidden' if hidden_sidlen else 'VISIBLE'}, "
        f"原姓名={'hidden' if hidden_origname else 'VISIBLE'}")

    # C7: freeze panes — at least 2 rows, 4 cols frozen (2 分)
    freeze_ok = False
    if ws.freeze_panes:
        try:
            fr, fc = coordinate_to_tuple(str(ws.freeze_panes))
            if (fr - 1) >= 2 and (fc - 1) >= 4:
                freeze_ok = True
        except Exception:
            pass
    add("C_格式与表格设置", "C7_freeze_panes", 2 if freeze_ok else 0,
        f"freeze={ws.freeze_panes}")

    # C8: 基本情况 wrap_text + vertical top (1 分)
    wrap_ok = True
    for r in range(3, min(last_row + 1, 50)):
        al = ws.cell(r, basic_col).alignment
        if al is None:
            continue
        if al.wrap_text is not True or al.vertical != "top":
            wrap_ok = False
            break
    add("C_格式与表格设置", "C8_basic_wrap", 1 if wrap_ok else 0,
        "wrap+top OK" if wrap_ok else "wrap/top incorrect")

    # C9: borders on header+data area (1 分)
    border_check = ws.cell(2, 1).border
    has_border = bool(border_check and (
        border_check.left.style or border_check.right.style
        or border_check.top.style or border_check.bottom.style))
    add("C_格式与表格设置", "C9_borders", 1 if has_border else 0,
        "borders detected" if has_border else "no borders on header row")

    # ==================================================================
    # B. 数据规范化 (20 分)
    # ==================================================================

    # B1: 学号+联系电话 text format (4 分)
    b1 = 0
    if ws.cell(3, sid_col).number_format == "@":
        b1 += 2
    if ws.cell(3, phone_col).number_format == "@":
        b1 += 2
    add("B_数据规范化", "B1_text_format", b1, f"text format score {b1}/4")

    # B2: 金额 — integer, no "元" (4 分)
    amt_issues = 0
    for r in range(3, last_row + 1):
        v = ws.cell(r, amt_col).value
        if v is None:
            continue
        sv = _s(v)
        if "元" in sv:
            amt_issues += 1
        elif isinstance(v, float) and v != int(v):
            amt_issues += 1
        elif isinstance(v, str):
            try:
                float(sv)  # should be numeric
            except ValueError:
                amt_issues += 1
    add("B_数据规范化", "B2_amount", 4 if amt_issues == 0 else 0,
        "amounts clean" if amt_issues == 0 else f"{amt_issues} amount issues")

    # B3: 在读学位 normalized (3 分)
    deg_issues = 0
    for r in range(3, last_row + 1):
        v = _s(ws.cell(r, deg_col).value)
        if v and v not in ("本科", "硕士", "博士"):
            deg_issues += 1
    add("B_数据规范化", "B3_degree", 3 if deg_issues == 0 else 0,
        "degrees OK" if deg_issues == 0 else f"{deg_issues} non-standard")

    # B4: 生源地 — no 省/市/自治区 (3 分)
    origin_issues = 0
    for r in range(3, last_row + 1):
        v = _s(ws.cell(r, origin_col).value)
        if any(suf in v for suf in ("省", "市", "自治区", "特别行政区")):
            origin_issues += 1
    add("B_数据规范化", "B4_origin", 3 if origin_issues == 0 else 0,
        "origins clean" if origin_issues == 0 else f"{origin_issues} with suffix")

    # B5: 民族 — no "族" (3 分)
    eth_issues = 0
    for r in range(3, last_row + 1):
        v = _s(ws.cell(r, eth_col).value)
        if "族" in v:
            eth_issues += 1
    add("B_数据规范化", "B5_ethnicity", 3 if eth_issues == 0 else 0,
        "ethnicity clean" if eth_issues == 0 else f"{eth_issues} with 族")

    # B6: 学院 full names match Sheet4 (3 分)
    acad_issues = []
    if acad_set:
        for r in range(3, last_row + 1):
            v = _s(ws.cell(r, acad_col).value)
            if v and v not in acad_set:
                acad_issues.append(v)
                if len(acad_issues) >= 5:
                    break
    if acad_set and not acad_issues:
        add("B_数据规范化", "B6_academy", 3, "all academy names match Sheet4")
    else:
        add("B_数据规范化", "B6_academy", 0,
            f"academy issues: {acad_issues[:3]}" if acad_issues
            else "no Sheet4 academy list available")

    # ==================================================================
    # A. 去重与排序 (25 分)
    # ==================================================================

    # A1: sequential numbering from 1 (5 分)
    seq_ok = bool(data_rows)
    for i, r in enumerate(data_rows, start=1):
        cv = ws.cell(r, 1).value
        if cv is None or int(cv) != i:
            seq_ok = False
            break
    add("A_去重与排序", "A1_sequential_num", 5 if seq_ok else 0,
        f"sequential 1..{len(data_rows)}" if seq_ok else "numbering incorrect")

    # A2: no duplicate students remain (8 分)
    seen_keys = set()
    remaining_dups = 0
    for r in data_rows:
        key = _dedup_key(
            ws.cell(r, sid_col).value,
            ws.cell(r, origname_col).value,
            ws.cell(r, phone_col).value,
        )
        if key in seen_keys:
            remaining_dups += 1
        seen_keys.add(key)
    add("A_去重与排序", "A2_no_duplicates", 8 if remaining_dups == 0 else 0,
        "no duplicates" if remaining_dups == 0
        else f"{remaining_dups} duplicates still present")

    # A3: sort order correct — academy → degree → review type → amount (10 分)
    sort_ok = False
    if academy_order and data_rows:
        prev_key = None
        sort_ok = True
        for r in data_rows:
            amt = _parse_int_amount(ws.cell(r, amt_col).value)
            cur_key = (
                acad_rank.get(_s(ws.cell(r, acad_col).value), 10**9),
                _degree_sort_key(ws.cell(r, deg_col).value),
                _review_sort_key(ws.cell(r, review_col).value),
                amt if amt >= 0 else 10**9,
            )
            if prev_key is not None and cur_key < prev_key:
                sort_ok = False
                break
            prev_key = cur_key
    add("A_去重与排序", "A3_sort_order", 10 if sort_ok else 0,
        "sort order correct" if sort_ok else "sort order wrong or cannot verify")

    # A4: duplicate students marked with red background (2 分)
    a4 = 0
    input_path = _locate_input_xlsx()
    if input_path and data_rows:
        try:
            wbi = openpyxl.load_workbook(input_path, data_only=True)
            wsi = wbi[wbi.sheetnames[0]]
            ih, ihmap = _headers_from_row(wsi, 2)

            def icol(name, fallback):
                if name in ihmap:
                    return ihmap[name]
                return column_index_from_string(fallback)

            i_sid = icol("学号", "D")
            i_name = icol("姓名", "E")
            i_phone = icol("联系电话", "P")

            # Count occurrences in original
            orig_counts = {}
            for r in range(3, wsi.max_row + 1):
                k = _dedup_key(
                    wsi.cell(r, i_sid).value,
                    wsi.cell(r, i_name).value,
                    wsi.cell(r, i_phone).value,
                )
                orig_counts[k] = orig_counts.get(k, 0) + 1
            duplicated_keys = {k for k, cnt in orig_counts.items() if cnt >= 2}

            # Map output rows by key
            out_map = {}
            for r in data_rows:
                k = _dedup_key(
                    ws.cell(r, sid_col).value,
                    ws.cell(r, origname_col).value,
                    ws.cell(r, phone_col).value,
                )
                out_map.setdefault(k, []).append(r)

            red_ok = True
            for dk in duplicated_keys:
                rows_for_key = out_map.get(dk, [])
                if len(rows_for_key) != 1:
                    red_ok = False
                    break
                rr = rows_for_key[0]
                # Check if any cell in the row is red
                found_red = False
                for cc in range(1, min(ws.max_column + 1, 25)):
                    if _is_red_fill(ws.cell(rr, cc)):
                        found_red = True
                        break
                if not found_red:
                    red_ok = False
                    break

            if red_ok and duplicated_keys:
                a4 = 2
        except Exception:
            pass
    add("A_去重与排序", "A4_dup_red_mark", a4,
        "red marking OK" if a4 == 2 else "red marking missing or incorrect")

    # ==================================================================
    # D. 公式与函数 (10 分)
    # ==================================================================

    # D1: LEN formula in 学号长度 column (4 分)
    sid_letter = get_column_letter(sid_col)
    len_total, len_formula_ok = 0, 0
    for r in range(3, last_row + 1):
        if not _s(ws.cell(r, sid_col).value):
            continue
        len_total += 1
        v = ws.cell(r, sidlen_col).value
        if isinstance(v, str) and v.startswith("="):
            norm = v.upper().replace(" ", "").replace("$", "")
            expected = f"=LEN({sid_letter}{r})".upper()
            if norm == expected:
                len_formula_ok += 1
    if len_total > 0 and len_formula_ok == len_total:
        add("D_公式与函数", "D1_len_formula", 4, "all LEN formulas correct")
    elif len_total > 0 and len_formula_ok / len_total >= 0.8:
        add("D_公式与函数", "D1_len_formula", 2,
            f"{len_formula_ok}/{len_total} LEN formulas correct")
    else:
        add("D_公式与函数", "D1_len_formula", 0,
            f"{len_formula_ok}/{len_total} LEN formulas")

    # D2: IF formula for name alignment (3 分)
    orig_letter = get_column_letter(origname_col)
    name_total, name_formula_ok = 0, 0
    for r in range(3, min(last_row + 1, 300)):
        v = ws.cell(r, name_col).value
        if v is None:
            continue
        name_total += 1
        if isinstance(v, str) and v.startswith("="):
            u = re.sub(r"\s+", "", v).replace("$", "").upper()
            if "IF(" in u and "LEN(" in u and f"{orig_letter}{r}" in u:
                name_formula_ok += 1
    if name_total > 0 and name_formula_ok / name_total >= 0.8:
        add("D_公式与函数", "D2_if_name", 3, "IF+LEN name formulas correct")
    elif name_formula_ok > 0:
        add("D_公式与函数", "D2_if_name", 1,
            f"{name_formula_ok}/{name_total} IF formulas found")
    else:
        add("D_公式与函数", "D2_if_name", 0, "no IF name formulas detected")

    # D3: Sheet3 COUNTIF and SUM formulas (3 分)
    d3 = 0
    if "Sheet3" in wb.sheetnames:
        ws3 = wb["Sheet3"]
        b1_val = ws3["B1"].value
        b2_val = ws3["B2"].value
        eth_letter = get_column_letter(eth_col)
        amt_letter = get_column_letter(amt_col)

        # Check COUNTIF for 汉族 count
        if isinstance(b1_val, str) and b1_val.startswith("="):
            u = re.sub(r"\s+", "", b1_val).replace("$", "")
            # Accept both full-range and column-range COUNTIF
            pat_full = re.compile(
                rf'^=COUNTIF\(汇总表!{eth_letter}\d+:{eth_letter}\d+,"汉"\)$', re.I)
            pat_col = re.compile(
                rf'^=COUNTIF\(汇总表!{eth_letter}:{eth_letter},"汉"\)$', re.I)
            if pat_full.match(u) or pat_col.match(u):
                d3 += 1

        # Check SUM for total amount
        if isinstance(b2_val, str) and b2_val.startswith("="):
            u = re.sub(r"\s+", "", b2_val).replace("$", "")
            pat_full = re.compile(
                rf"^=SUM\(汇总表!{amt_letter}\d+:{amt_letter}\d+\)$", re.I)
            pat_col = re.compile(
                rf"^=SUM\(汇总表!{amt_letter}:{amt_letter}\)$", re.I)
            if pat_full.match(u) or pat_col.match(u):
                d3 += 1

    if d3 == 2:
        add("D_公式与函数", "D3_sheet3_formulas", 3, "COUNTIF + SUM both correct")
    elif d3 == 1:
        add("D_公式与函数", "D3_sheet3_formulas", 1, "only one formula correct")
    else:
        add("D_公式与函数", "D3_sheet3_formulas", 0, "Sheet3 formulas missing/wrong")

    # ==================================================================
    # E. Sheet2/Sheet3 (10 分)
    # ==================================================================

    # E1: 打印版 visible columns (5 分)
    e1 = 0
    if "打印版" in wb.sheetnames:
        ws2 = wb["打印版"]
        vis = []
        for c in range(1, ws2.max_column + 1):
            letter = get_column_letter(c)
            dim = ws2.column_dimensions.get(letter)
            if dim and dim.hidden:
                continue
            v = _s(ws2.cell(2, c).value)
            if v:
                vis.append(v)
        target_vis = ["序号", "学号", "姓名", "学院", "金额",
                      "在读学位", "民族", "评审类型"]
        if vis == target_vis:
            e1 = 5
        elif set(target_vis).issubset(set(vis)):
            e1 = 2  # has extra visible columns
    add("E_Sheet2_Sheet3", "E1_print_cols", e1,
        "visible columns correct" if e1 == 5 else f"visible cols score {e1}/5")

    # E2: print settings — landscape + centered/fit (3 分)
    e2 = 0
    if "打印版" in wb.sheetnames:
        ws2 = wb["打印版"]
        ps = ws2.page_setup
        po = getattr(ws2, "print_options", None)
        if getattr(ps, "orientation", None) == "landscape":
            e2 += 1
        if po and getattr(po, "horizontalCentered", False):
            e2 += 1
        if getattr(ps, "fitToWidth", None) in (1, "1"):
            e2 += 1
        e2 = min(e2, 3)
    add("E_Sheet2_Sheet3", "E2_print_settings", e2, f"print settings {e2}/3")

    # E3: Sheet3 labels (2 分)
    e3 = 0
    if "Sheet3" in wb.sheetnames:
        ws3 = wb["Sheet3"]
        if (_s(ws3["A1"].value) == "汉族同学人数"
                and _s(ws3["A2"].value) == "奖学金总金额"):
            e3 = 2
    add("E_Sheet2_Sheet3", "E3_sheet3_labels", e3,
        "labels correct" if e3 == 2 else "labels missing/wrong")

    # ==================================================================
    # Final score
    # ==================================================================
    total = sum(s["score"] for s in sections.values())
    total = max(0, min(100, total))

    return total, _build_report(sections, total)


def _build_report(sections: dict, total: int) -> Dict[str, Any]:
    """Build the full report dict from section data."""
    scores_summary = {}
    details = {}
    for key, sec in sections.items():
        scores_summary[key] = sec["score"]
        details[key] = sec["details"]
    return {"total": total, "scores": scores_summary, "details": details}


# ---------------------------------------------------------------------------
# print_report
# ---------------------------------------------------------------------------

def print_report(score: int, report: Dict[str, Any]) -> None:
    """Print a formatted evaluation report."""
    print("=" * 70)
    print("EXCEL汇总表整理 — 评分报告")
    print("=" * 70)
    print(f"\n总分: {score}/100\n")

    dim_labels = {
        "F_文件完整性": ("F. 文件完整性", 5),
        "A_去重与排序": ("A. 去重与排序", 25),
        "B_数据规范化": ("B. 数据规范化", 20),
        "C_格式与表格设置": ("C. 格式与表格设置", 30),
        "D_公式与函数": ("D. 公式与函数", 10),
        "E_Sheet2_Sheet3": ("E. Sheet2/Sheet3", 10),
    }

    scores = report.get("scores", {})
    details = report.get("details", {})

    print("分项得分:")
    for key, (label, max_pts) in dim_labels.items():
        pts = scores.get(key, 0)
        print(f"  {label}: {pts}/{max_pts}")

    print()

    for key, (label, _) in dim_labels.items():
        section = details.get(key, {})
        if section:
            print(f"--- {label} ---")
            for k, v in section.items():
                print(f"  {k}: {v}")
            print()

    print("=" * 70)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        test_dir = sys.argv[1]
    else:
        test_dir = os.path.join(os.path.dirname(__file__), "..", "gpt-5", "attempt_1")
    test_dir = os.path.abspath(test_dir)
    if os.path.exists(test_dir):
        print(f"Evaluating: {test_dir}\n")
        s, r = evaluate(test_dir)
        print_report(s, r)
    else:
        print(f"Directory not found: {test_dir}")
    sys.exit(0)
